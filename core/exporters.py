import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font, Side, Border
from openpyxl.utils import get_column_letter
from io import BytesIO
import zipfile
from PIL import Image, ImageDraw, ImageFont
import numpy as np

def export_to_excel_multi(code_blocks, img_blocks, alpha_blocks, rgb_to_code, block_size, global_counts, grid_shape):
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment, Font, Side, Border
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = openpyxl.Workbook()
    recap_ws = wb.active
    recap_ws.title = "Récap"

    # Récap des couleurs
    if global_counts:
        recap_ws.cell(row=1, column=1, value="Code couleur")
        recap_ws.cell(row=1, column=2, value="Quantité")
        recap_ws.cell(row=1, column=1).font = Font(bold=True)
        recap_ws.cell(row=1, column=2).font = Font(bold=True)
        for idx, (code, qty) in enumerate(global_counts, start=2):
            recap_ws.cell(row=idx, column=1, value=code)
            recap_ws.cell(row=idx, column=2, value=qty)

    # Bordures
    thin_black = Side(border_style="thin", color="000000")
    thick_black = Side(border_style="medium", color="000000")  # ~3px
    thick_red = Side(border_style="medium", color="FF0000")    # ~3px

    # Onglet par bloc
    for entry in code_blocks:
        if isinstance(entry, tuple) and len(entry) == 3:
            i, j, cblock = entry
            sheet_name = f"Bloc_{i//block_size}_{j//block_size}"
            ws = wb.create_sheet(title=sheet_name)
            h, w = cblock.shape
            for x in range(h):
                for y in range(w):
                    code = str(cblock[x, y])
                    cell = ws.cell(row=x+1, column=y+1)
                    cell.value = code
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # Coloration si code connu
                    if code and code in rgb_to_code:
                        # Accepts dict or tuple for rgb_to_code
                        rgb = rgb_to_code[code]['rgb'] if isinstance(rgb_to_code[code], dict) else rgb_to_code[code]
                        try:
                            rgb = tuple(int(float(xx)) for xx in rgb)
                            hex_color = '{0:02X}{1:02X}{2:02X}'.format(*rgb)
                        except Exception:
                            hex_color = "FFFFFF"
                        cell.fill = PatternFill(
                            start_color=hex_color,
                            end_color=hex_color,
                            fill_type='solid'
                        )
                        luminance = 0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2]
                        font_color = "FFFFFF" if luminance < 128 else "000000"
                        cell.font = Font(size=10, color=font_color)
                    elif not code:
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                    # ---- Bordures ----
                    left_side = thin_black
                    top_side = thin_black
                    right_side = thin_black
                    bottom_side = thin_black

                    # Grosse bordure noire chaque 10 cellules
                    if (x+1) % 10 == 0:
                        bottom_side = thick_red
                    if (y+1) % 10 == 0:
                        right_side = thick_red
                    

                    cell.border = Border(
                        left=left_side,
                        top=top_side,
                        right=right_side,
                        bottom=bottom_side
                    )


            # Ajustement tailles
            col_width = 5
            row_height = 24
            for col in range(1, w + 1):
                ws.column_dimensions[get_column_letter(col)].width = col_width
            for row in range(1, h + 1):
                ws.row_dimensions[row].height = row_height

    # Supprime la feuille par défaut si vide
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def make_bead_grid_image(
    code_grid, color_grid, alpha_grid, rgb_to_code, bead_size=80, margin=10, font_path=None,
    block_name=None, block_size=None
):
    from PIL import Image, ImageDraw, ImageFont
    h, w = code_grid.shape
    width_px = w * bead_size + 2 * margin
    height_px = h * bead_size + 2 * margin
    img = Image.new("RGB", (width_px, height_px), "white")
    draw = ImageDraw.Draw(img)
    if font_path is not None:
        font = ImageFont.truetype(font_path, int(bead_size/2))
    else:
        try:
            font = ImageFont.truetype("arial.ttf", int(bead_size/2.7))
        except:
            font = ImageFont.load_default()
    for i in range(h):
        for j in range(w):
            if code_grid[i, j] == '' or (alpha_grid is not None and alpha_grid[i, j] == 0):
                continue
            color = tuple(color_grid[i, j])
            code = str(code_grid[i, j])
            x = margin + j * bead_size
            y = margin + i * bead_size
            bbox = [x, y, x+bead_size, y+bead_size]
            draw.ellipse(bbox, fill=color, outline="grey", width=2)
            luminance = 0.299 * color[0] + 0.587 * color[1] + 0.114 * color[2]
            txt_color = "white" if luminance < 128 else "black"
            try:
                bbox_txt = font.getbbox(code)
            except AttributeError:
                bbox_txt = draw.textbbox((0, 0), code, font=font)
            wtxt = bbox_txt[2] - bbox_txt[0]
            htxt = bbox_txt[3] - bbox_txt[1]
            offset = 1
            outline_color = "black" if txt_color == "white" else "white"
            for dx in [-offset, 0, offset]:
                for dy in [-offset, 0, offset]:
                    if dx == 0 and dy == 0:
                        continue
                    draw.text(
                        (x + bead_size/2 - wtxt/2 + dx, y + bead_size/2 - htxt/2 + dy),
                        code, fill=outline_color, font=font
                    )
            draw.text(
                (x + bead_size/2 - wtxt/2, y + bead_size/2 - htxt/2),
                code,
                fill=txt_color,
                font=font
            )
    if block_name:
        label = f"Bloc {block_name}"
        draw.text((margin, 2), label, fill="red", font=font)

    # --- Grille grise fine entre chaque perle ---
    gray = (180, 180, 180)
    for x in range(1, w):
        px = margin + x * bead_size
        draw.line([(px, margin), (px, margin + h * bead_size)], fill=gray, width=1)
    for y in range(1, h):
        py = margin + y * bead_size
        draw.line([(margin, py), (margin + w * bead_size, py)], fill=gray, width=1)

    # --- Lignes noires 2px tous les 10 perles ---
    for x in range(10, w, 10):
        px = margin + x * bead_size
        draw.line([(px, margin), (px, margin + h * bead_size)], fill=(0, 0, 0), width=2)
    for y in range(10, h, 10):
        py = margin + y * bead_size
        draw.line([(margin, py), (margin + w * bead_size, py)], fill=(0, 0, 0), width=2)

    # --- Lignes rouges 4px tous les block_size ---
    if block_size and block_size > 1:
        for x in range(block_size, w, block_size):
            px = margin + x * bead_size
            draw.line([(px, margin), (px, margin + h * bead_size)], fill=(255, 0, 0), width=4)
        for y in range(block_size, h, block_size):
            py = margin + y * bead_size
            draw.line([(margin, py), (margin + w * bead_size, py)], fill=(255, 0, 0), width=4)

    return img

def export_png_and_zip(code_grid, color_grid, alpha_grid, rgb_to_code, block_size, font_path=None):
    from io import BytesIO
    import zipfile
    from core.image_processing import split_grid

    h, w = code_grid.shape
    img_full = make_bead_grid_image(
        code_grid, color_grid, alpha_grid, rgb_to_code,
        bead_size=80, margin=10, font_path=font_path,
        block_size=block_size
    )
    buf_full = BytesIO()
    img_full.save(buf_full, format='PNG')
    buf_full.seek(0)

    code_blocks = split_grid(code_grid, block_size)
    color_blocks = split_grid(color_grid, block_size)
    alpha_blocks = split_grid(alpha_grid, block_size)
    block_imgs = []
    for (i, j, cblock), (_, _, clblock), (_, _, alblock) in zip(code_blocks, color_blocks, alpha_blocks):
        block_name = f"{i//block_size},{j//block_size}"
        img_block = make_bead_grid_image(
            cblock, clblock, alblock, rgb_to_code,
            bead_size=80, margin=10, block_name=block_name,
            block_size=block_size
        )
        buf_block = BytesIO()
        img_block.save(buf_block, format='PNG')
        buf_block.seek(0)
        block_imgs.append((block_name, buf_block))

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zipf:
        zipf.writestr('image_complete.png', buf_full.read())
        for block_name, buf in block_imgs:
            zipf.writestr(f'bloc_{block_name}.png', buf.read())
    zip_buffer.seek(0)
    return zip_buffer
