import streamlit as st
from PIL import Image, ImageDraw, ImageFont
import math
import pandas as pd
import numpy as np
import base64
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from collections import Counter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import zipfile, tempfile

# Fonctions utilitaires
def draw_block_grid(image, block_size, color=(255, 0, 0), width=1):
    """Dessine des lignes de séparation sur l'image pour visualiser les blocs."""
    draw = ImageDraw.Draw(image)
    img_w, img_h = image.size
    # Lignes verticales
    for x in range(block_size, img_w, block_size):
        draw.line([(x, 0), (x, img_h)], fill=color, width=width)
    # Lignes horizontales
    for y in range(block_size, img_h, block_size):
        draw.line([(0, y), (img_w, y)], fill=color, width=width)
    return image

def find_closest_color(pixel, palette):
    diffs = palette - pixel
    dist = np.sum(diffs**2, axis=1)
    idx = np.argmin(dist)
    return palette[idx]

def pil_to_base64(img):
    buf = BytesIO()
    img.save(buf, format="PNG")
    b64 = base64.b64encode(buf.getvalue()).decode()
    return b64

def export_to_excel(code_grid, result_img, rgb_to_code, alpha_arr, filename="export_perler.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grille"

    h, w = code_grid.shape

    for i in range(h):
        for j in range(w):
            code = code_grid[i, j]
            if code == '' or code is None or (alpha_arr is not None and alpha_arr[i, j] == 0):
                continue  # Pas de perle ici (transparent)
            cell = ws.cell(row=i+1, column=j+1)
            cell.value = code
            # Couleur de fond (ARGB)
            rgb = result_img[i, j]
            hex_color = "FF{:02X}{:02X}{:02X}".format(*rgb)
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            luminance = 0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2]
            font_color = "FFFFFF" if luminance < 128 else "000000"
            cell.font = Font(color=font_color, size=8, bold=False)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(j+1)].width = 4

    code_list = code_grid.flatten()
    counts = Counter([code for code in code_list if code != '' and code is not None])
    start_row = h + 3
    ws.cell(row=start_row, column=1).value = "Code couleur"
    ws.cell(row=start_row, column=2).value = "Quantité"
    for idx, (code, count) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), 1):
        ws.cell(row=start_row + idx, column=1).value = code
        ws.cell(row=start_row + idx, column=2).value = count

    file = BytesIO()
    wb.save(file)
    file.seek(0)
    return file

def export_to_excel_multi(code_blocks, img_blocks, alpha_blocks, rgb_to_code, block_size, global_counts, grid_shape):
    wb = openpyxl.Workbook()
    # Crée la feuille Récap d'abord
    recap_ws = wb.active
    recap_ws.title = "Récap"

    # 1. Tableau de placement des blocs
    n_blocks_rows = math.ceil(grid_shape[0] / block_size)
    n_blocks_cols = math.ceil(grid_shape[1] / block_size)
    recap_ws.cell(row=1, column=1).value = "Disposition des blocs"
    for i in range(n_blocks_rows):
        for j in range(n_blocks_cols):
            recap_ws.cell(row=i+2, column=j+2).value = f"Bloc {i},{j}"
            recap_ws.cell(row=i+2, column=j+2).alignment = Alignment(horizontal="center", vertical="center")
            recap_ws.cell(row=i+2, column=j+2).font = Font(bold=True)
            recap_ws.cell(row=i+2, column=j+2).fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")

    # 2. Tableau récap total
    start_row = n_blocks_rows + 4
    recap_ws.cell(row=start_row, column=1).value = "Code couleur"
    recap_ws.cell(row=start_row, column=2).value = "Quantité"
    recap_ws.cell(row=start_row, column=1).font = Font(bold=True)
    recap_ws.cell(row=start_row, column=2).font = Font(bold=True)
    for idx, (code, count) in enumerate(global_counts, 1):
        recap_ws.cell(row=start_row + idx, column=1).value = code
        recap_ws.cell(row=start_row + idx, column=2).value = count

    # 3. Les feuilles de blocs
    for idx, ((i, j, cblock), (_, _, iblock), (_, _, ablock)) in enumerate(zip(code_blocks, img_blocks, alpha_blocks)):
        h, w = cblock.shape
        ws = wb.create_sheet(title=f"Bloc {i//block_size},{j//block_size}")
        for x in range(h):
            for y in range(w):
                code = cblock[x, y]
                if code == '' or code is None or (ablock is not None and ablock[x, y] == 0):
                    continue
                cell = ws.cell(row=x+1, column=y+1)
                cell.value = code
                hex_color = "FF{:02X}{:02X}{:02X}".format(*iblock[x, y])
                cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                luminance = 0.299 * iblock[x, y][0] + 0.587 * iblock[x, y][1] + 0.114 * iblock[x, y][2]
                font_color = "FFFFFF" if luminance < 128 else "000000"
                cell.font = Font(color=font_color, size=8, bold=False)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[get_column_letter(y+1)].width = 4

        # Tableau récap du bloc (en bas)
        code_list = cblock.flatten()
        counts = Counter([code for code in code_list if code != '' and code is not None])
        start_row_bloc = h + 3
        ws.cell(row=start_row_bloc, column=1).value = "Code couleur"
        ws.cell(row=start_row_bloc, column=2).value = "Quantité"
        ws.cell(row=start_row_bloc, column=1).font = Font(bold=True)
        ws.cell(row=start_row_bloc, column=2).font = Font(bold=True)
        for rec_idx, (code, count) in enumerate(sorted(counts.items(), key=lambda x: -x[1]), 1):
            ws.cell(row=start_row_bloc + rec_idx, column=1).value = code
            ws.cell(row=start_row_bloc + rec_idx, column=2).value = count

    file = BytesIO()
    wb.save(file)
    file.seek(0)
    return file


def export_to_pdf(code_grid, result_img, counts, filename="export_perler.pdf"):
    buf = BytesIO()
    pagesize = landscape(A4) if code_grid.shape[1] > code_grid.shape[0] else A4
    doc = SimpleDocTemplate(buf, pagesize=pagesize)
    elements = []
    styleSheet = getSampleStyleSheet()

    h, w = code_grid.shape
    elements.append(Paragraph("Grille Perler Art", styleSheet["Title"]))
    elements.append(Spacer(1, 8))

    cell_size = 12
    max_dim = 60
    disp_h, disp_w = min(h, max_dim), min(w, max_dim)
    table_data = []
    for i in range(disp_h):
        row = []
        for j in range(disp_w):
            code = code_grid[i, j]
            txt = "" if code == '' or code is None else code
            row.append(txt)
        table_data.append(row)
    tbl = Table(table_data, colWidths=cell_size, rowHeights=cell_size)
    style = TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 7),
        ('GRID', (0,0), (-1,-1), 0.1, colors.grey),
    ])
    for i in range(disp_h):
        for j in range(disp_w):
            code = code_grid[i, j]
            if code != '' and code is not None:
                rgb = result_img[i, j]
                bg = colors.Color(rgb[0]/255, rgb[1]/255, rgb[2]/255)
                style.add('BACKGROUND', (j, i), (j, i), bg)
    tbl.setStyle(style)
    elements.append(tbl)

    elements.append(Spacer(1, 16))
    elements.append(Paragraph("Récapitulatif des perles nécessaires :", styleSheet['Heading2']))
    recap_data = [["Code couleur", "Quantité"]]
    for code, count in counts:
        recap_data.append([code, count])
    recap_tbl = Table(recap_data, colWidths=[50, 50])
    recap_tbl.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('GRID', (0,0), (-1,-1), 0.2, colors.black)
    ]))
    elements.append(recap_tbl)

    doc.build(elements)
    buf.seek(0)
    return buf

def export_to_pdf_multi(code_blocks, img_blocks, block_size):
    buf = BytesIO()
    elements = []
    styleSheet = getSampleStyleSheet()
    cell_size = 12
    max_dim = 60

    from reportlab.platypus import PageBreak

    for idx, ((i, j, cblock), (_, _, iblock)) in enumerate(zip(code_blocks, img_blocks)):
        h, w = cblock.shape
        disp_h, disp_w = min(h, max_dim), min(w, max_dim)
        elements.append(Paragraph(f"Bloc {i}:{i+h}, {j}:{j+w}", styleSheet["Title"]))
        elements.append(Spacer(1, 8))
        table_data = []
        for x in range(disp_h):
            row = []
            for y in range(disp_w):
                code = cblock[x, y]
                txt = "" if code == '' or code is None else code
                row.append(txt)
            table_data.append(row)
        tbl = Table(table_data, colWidths=cell_size, rowHeights=cell_size)
        style = TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,0), (-1,-1), 7),
            ('GRID', (0,0), (-1,-1), 0.1, colors.grey),
        ])
        for x in range(disp_h):
            for y in range(disp_w):
                code = cblock[x, y]
                if code != '' and code is not None:
                    rgb = iblock[x, y]
                    bg = colors.Color(rgb[0]/255, rgb[1]/255, rgb[2]/255)
                    style.add('BACKGROUND', (y, x), (y, x), bg)
        tbl.setStyle(style)
        elements.append(tbl)
        elements.append(PageBreak())
    # Remove the last page break
    if elements and isinstance(elements[-1], PageBreak):
        elements = elements[:-1]
    pagesize = landscape(A4)
    doc = SimpleDocTemplate(buf, pagesize=pagesize)
    doc.build(elements)
    buf.seek(0)
    return buf


def split_grid(arr, size):
    h, w = arr.shape[:2]
    blocks = []
    for i in range(0, h, size):
        for j in range(0, w, size):
            blocks.append((i, j, arr[i:i+size, j:j+size]))
    return blocks

def export_png_and_zip(code_grid, color_grid, alpha_grid, rgb_to_code, block_size, font_path=None):
    from io import BytesIO
    h, w = code_grid.shape

    # 1. Image complète avec traits rouges
    img_full = make_bead_grid_image(code_grid, color_grid, alpha_grid, rgb_to_code, bead_size=80, margin=10,
                                    font_path=font_path, block_size=block_size, grid_color=(255,0,0))
    buf_full = BytesIO()
    img_full.save(buf_full, format='PNG')
    buf_full.seek(0)

    # 2. Création des blocs
    code_blocks = split_grid(code_grid, block_size)
    color_blocks = split_grid(color_grid, block_size)
    alpha_blocks = split_grid(alpha_grid, block_size)
    block_imgs = []
    for (i, j, cblock), (_, _, clblock), (_, _, alblock) in zip(code_blocks, color_blocks, alpha_blocks):
        block_name = f"{i//block_size},{j//block_size}"
        img_block = make_bead_grid_image(cblock, clblock, alblock, rgb_to_code, bead_size=80, margin=10, block_name=block_name)
        buf_block = BytesIO()
        img_block.save(buf_block, format='PNG')
        buf_block.seek(0)
        block_imgs.append((block_name, buf_block))

    # 3. Crée le ZIP
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zipf:
        # Ajoute l’image complète
        zipf.writestr('image_complete.png', buf_full.read())
        # Ajoute chaque bloc
        for block_name, buf in block_imgs:
            zipf.writestr(f'bloc_{block_name}.png', buf.read())
    zip_buffer.seek(0)
    return zip_buffer

def make_bead_grid_image(code_grid, color_grid, alpha_grid, rgb_to_code, bead_size=80, margin=10, font_path=None, block_name=None, block_size=None, grid_color=(255,0,0)):
    h, w = code_grid.shape
    width_px = w * bead_size + 2 * margin
    height_px = h * bead_size + 2 * margin
    img = Image.new("RGB", (width_px, height_px), "white")
    draw = ImageDraw.Draw(img)
    
    # Police pour les codes couleur
    if font_path is not None:
        font = ImageFont.truetype(font_path, int(bead_size/2))
    else:
        try:
            font = ImageFont.truetype("arial.ttf", int(bead_size/2.7))
        except:
            font = ImageFont.load_default()


    # Cercles par perle
    for i in range(h):
        for j in range(w):
            if code_grid[i, j] == '' or (alpha_grid is not None and alpha_grid[i, j] == 0):
                continue
            color = tuple(color_grid[i, j])
            code = str(code_grid[i, j])
            x = margin + j * bead_size
            y = margin + i * bead_size
            bbox = [x, y, x+bead_size, y+bead_size]
            draw.ellipse(bbox, fill=color, outline="grey", width=2)

            # Couleur du texte selon la luminosité
            luminance = 0.299 * color[0] + 0.587 * color[1] + 0.114 * color[2]
            txt_color = "white" if luminance < 128 else "black"

            # Centrage texte
            try:
                bbox_txt = font.getbbox(code)
            except AttributeError:
                bbox_txt = draw.textbbox((0, 0), code, font=font)
            wtxt = bbox_txt[2] - bbox_txt[0]
            htxt = bbox_txt[3] - bbox_txt[1]
            offset = 1
            # Contour blanc si txt_color est noir, ou noir sinon
            outline_color = "black" if txt_color == "white" else "white"
            for dx in [-offset, 0, offset]:
                for dy in [-offset, 0, offset]:
                    if dx == 0 and dy == 0:
                        continue
                    draw.text(
                        (x + bead_size/2 - wtxt/2 + dx, y + bead_size/2 - htxt/2 + dy),
                        code, fill=outline_color, font=font
                    )

            draw.text(
                (x + bead_size/2 - wtxt/2, y + bead_size/2 - htxt/2),
                code,
                fill=txt_color,
                font=font
            )

    
    # Ajoute nom du bloc si précisé
    if block_name:
        label = f"Bloc {block_name}"
        draw.text((margin, 2), label, fill="red", font=font)
    
    # Option: traits rouges pour la grille des blocs (pour l’image complète uniquement)
    if block_size and grid_color:
        # Traits verticaux
        for x in range(block_size, w, block_size):
            px = margin + x * bead_size
            draw.line([(px, margin), (px, margin + h * bead_size)], fill=grid_color, width=1)
        # Traits horizontaux
        for y in range(block_size, h, block_size):
            py = margin + y * bead_size
            draw.line([(margin, py), (margin + w * bead_size, py)], fill=grid_color, width=1)
    
    return img



# --- App Streamlit ---
st.title("Perler Art Pixelizer")

palettes = {
    "Artkal Mini C": "artkal_c_mini.csv",
    "Hama": "hama.csv",
    "Importer mon CSV...": None
}
palette_choice = st.selectbox("Choisir une palette de couleurs", list(palettes.keys()))

if palette_choice != "Importer mon CSV...":
    palette_filename = palettes[palette_choice]
    palette_df = pd.read_csv(palette_filename)
else:
    palette_df = None

if palette_df is not None:
    valid_palette = palette_df.dropna(subset=['R', 'G', 'B'])
    rgb_palette = []
    for i, row in valid_palette.iterrows():
        try:
            rgb_palette.append([int(row['R']), int(row['G']), int(row['B'])])
        except:
            continue
    rgb_palette = np.array(rgb_palette)

    st.subheader("Aperçu de la palette")
    cell_size = 32
    colors_html = ""
    for idx, row in palette_df.iterrows():
        try:
            color = '#{:02X}{:02X}{:02X}'.format(int(row['R']), int(row['G']), int(row['B']))
        except:
            continue
        code_str = str(row["Code"])
        txt_col = "#fff" if int(row['R'])*0.299 + int(row['G'])*0.587 + int(row['B'])*0.114 < 128 else "#000"
        colors_html += f"""
            <div style='
                width:{cell_size}px; height:{cell_size}px; background:{color};
                display:inline-block; margin:5px 4px 0 0; border-radius:8px; border:1.5px solid #888; vertical-align:top; position:relative; text-align:center;'>
                <span style='position:absolute; left:0; right:0; bottom:2px; font-size:0.7em; color:{txt_col}; font-family:monospace;'></span>
            </div>
        """
    st.markdown(colors_html, unsafe_allow_html=True)




uploaded_file = st.file_uploader("Uploader une image", type=['jpg', 'jpeg', 'png'])

if uploaded_file is not None:
    img = Image.open(uploaded_file)
    if img.mode != "RGBA":
        img = img.convert("RGBA")

    # Réinitialise le slider si nouvelle image
    if "last_uploaded_filename" not in st.session_state or uploaded_file.name != st.session_state.last_uploaded_filename:
        st.session_state.slider_width = img.width
        st.session_state.initial_width = img.width
        st.session_state.last_uploaded_filename = uploaded_file.name

    b64_original = pil_to_base64(img)
    st.markdown(
        f"""
        <div style='width:100%; max-width:500px; margin:auto;'>
            <img src='data:image/png;base64,{b64_original}'
            style='width:100%; height:auto; image-rendering:pixelated; border:1px solid #888;'/>
        </div>
        """,
        unsafe_allow_html=True
    )
    st.success("Image chargée !")

    # Pixel perfect, taille réelle de l'image originale
    b64_orig = pil_to_base64(img)
    width_orig, height_orig = img.size
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        if st.button("Format initial"):
            st.session_state.slider_width = st.session_state.initial_width
    with col2:
        if st.button("1/2"):
            st.session_state.slider_width = max(8, st.session_state.initial_width // 2)
    with col3:
        if st.button("1/4"):
            st.session_state.slider_width = max(8, st.session_state.initial_width // 4)
    with col4:
        if st.button("1/8"):
            st.session_state.slider_width = max(8, st.session_state.initial_width // 8)

    width = st.slider(
        "Largeur finale en pixels",
        min_value=8,
        max_value=img.width,
        value=st.session_state.slider_width,
        step=1,
        key="slider_width"
    )

    ratio = img.height / img.width
    new_height = int(width * ratio)
    st.write(f"Nouvelle taille : {width} x {new_height}")

    resized_img = img.resize((width, new_height), Image.NEAREST)
    b64_resized = pil_to_base64(resized_img)
    st.markdown(
        f"""
        <div style='width:100%; max-width:500px; margin:auto;'>
            <img src='data:image/png;base64,{b64_resized}'
            style='width:100%; height:auto; image-rendering:pixelated; border:1px solid #888;'/>
        </div>
        """,
        unsafe_allow_html=True
    )

    # -- Sélection taille des blocs --
    st.markdown("### Export Excel par blocs")
    colb1, colb2, colb3, colb4, colb5 = st.columns([1,1,1,2,1])
    if "block_size" not in st.session_state:
        st.session_state.block_size = 50
    with colb1:
        if st.button("19 x 19"):
            st.session_state.block_size = 19
    with colb2:
        if st.button("25 x 25"):
            st.session_state.block_size = 25
    with colb3:
        if st.button("50 x 50"):
            st.session_state.block_size = 50
    with colb4:
        custom_size = st.number_input("Taille custom", min_value=5, max_value=100, value=st.session_state.block_size, step=1)
        if custom_size != st.session_state.block_size:
            st.session_state.block_size = custom_size

    block_size = st.session_state.block_size
    st.write(f"Taille des blocs actuelle : {block_size} x {block_size}")

    # Bouton centré
    col_left, col_center, col_right = st.columns([2,1,2])
    with col_center:
        convert_clicked = st.button("Convertir l'image avec la palette")

    if convert_clicked:
        with st.spinner("Conversion en cours..."):
            arr = np.array(resized_img)
            if arr.shape[2] == 4:  # Image RGBA
                rgb_arr = arr[..., :3]
                alpha_arr = arr[..., 3]
            else:
                rgb_arr = arr
                alpha_arr = np.ones(arr.shape[:2], dtype=np.uint8) * 255

            result_img = np.zeros_like(rgb_arr)
            code_grid = np.empty(rgb_arr.shape[:2], dtype=object)
            # RGB->Code
            rgb_to_code = {}
            for i, row in valid_palette.iterrows():
                try:
                    rgb_tuple = (int(row['R']), int(row['G']), int(row['B']))
                    rgb_to_code[rgb_tuple] = row['Code']
                except:
                    continue

            for i in range(rgb_arr.shape[0]):
                for j in range(rgb_arr.shape[1]):
                    if alpha_arr[i, j] == 0:
                        code_grid[i, j] = ''
                        result_img[i, j] = [255, 255, 255]  # Option: blanc pour les transparents
                    else:
                        color = find_closest_color(rgb_arr[i, j], rgb_palette)
                        result_img[i, j] = color
                        code_grid[i, j] = rgb_to_code.get(tuple(color), '??')

            result_img_pil = Image.fromarray(result_img.astype('uint8'))

        st.success("Conversion terminée !")

        # Affichage image convertie AVEC lignes rouges de séparation de blocs
        block_size = st.session_state.block_size

        img_with_grid = result_img_pil.copy()
        img_with_grid = draw_block_grid(img_with_grid, block_size, color=(255,0,0), width=1)  # width=2 plus visible
        b64_img_grid = pil_to_base64(img_with_grid)
        st.markdown(
            f"""
            <div style='width:100%; max-width:500px; margin:auto;'>
                <img src='data:image/png;base64,{b64_img_grid}'
                style='width:100%; height:auto; image-rendering:pixelated; border:2px solid #888;'/>
            </div>
            <div style='font-size:12px; color:#d00;'>Aperçu converti – blocs visibles ({block_size} x {block_size})</div>
            """,
            unsafe_allow_html=True
        )

        # Tableau récap Streamlit
        code_list = code_grid.flatten()
        counts = Counter([code for code in code_list if code != '' and code is not None])
        sorted_counts = sorted(counts.items(), key=lambda x: -x[1])
        df_récap = pd.DataFrame(sorted_counts, columns=['Code couleur', 'Quantité'])
        # Calcul du nombre de blocs
        n_blocks_rows = math.ceil(code_grid.shape[0] / block_size)
        n_blocks_cols = math.ceil(code_grid.shape[1] / block_size)
        n_blocks_total = n_blocks_rows * n_blocks_cols

        # Affichage du nombre total de blocs juste avant le tableau récapitulatif
        st.markdown(f"**Nombre total de blocs : {n_blocks_total}**")
        st.subheader("Récapitulatif des perles nécessaires :")
        st.dataframe(df_récap, hide_index=True)

        n_pixels = code_grid.shape[0] * code_grid.shape[1]
        block_size = st.session_state.block_size

        if n_pixels <= 10000:
            st.subheader("Exporter toute l'image :")
            # Export Excel (tout en un)
            excel_file = export_to_excel(code_grid, result_img, rgb_to_code, alpha_arr)
            st.download_button(
                label="Télécharger l'export Excel",
                data=excel_file,
                file_name="perler_export.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            # Export PDF (tout en un)
            # pdf_file = export_to_pdf(code_grid, result_img, sorted_counts)
            # st.download_button(
            #     label="Télécharger l'export PDF",
            #     data=pdf_file,
            #     file_name="perler_export.pdf",
            #     mime="application/pdf"
            # )


            #Export en PNG
            img_png = make_bead_grid_image(code_grid, result_img, alpha_arr, rgb_to_code, bead_size=80, margin=10, block_size=block_size, grid_color=(255,0,0))
            buf_png = BytesIO()
            img_png.save(buf_png, format='PNG')
            buf_png.seek(0)
            st.download_button(
                label="Télécharger le PNG grille complète",
                data=buf_png,
                file_name="perler_complete.png",
                mime="image/png"
            )
        else:
            # Split en blocs
            code_blocks = split_grid(code_grid, block_size)
            img_blocks = split_grid(result_img, block_size)
            alpha_blocks = split_grid(alpha_arr, block_size)

            st.subheader("Exporter par blocs (un seul fichier) :")

            # Excel multi-sheets
            # Récap global toutes couleurs
            code_list_all = code_grid.flatten()
            counts_all = Counter([code for code in code_list_all if code != '' and code is not None])
            sorted_counts_all = sorted(counts_all.items(), key=lambda x: -x[1])
            with st.spinner("Préparation de l'export Excel…"):
                excel_file = export_to_excel_multi(
                    code_blocks, img_blocks, alpha_blocks, rgb_to_code, block_size, 
                    sorted_counts_all, code_grid.shape
                )
            st.download_button(
                label="Télécharger le fichier Excel (blocs)",
                data=excel_file,
                file_name="perler_export_blocs.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # PDF multipages
            # pdf_file = export_to_pdf_multi(code_blocks, img_blocks, block_size)
            # st.download_button(
            #     label="Télécharger le fichier PDF (blocs)",
            #     data=pdf_file,
            #     file_name="perler_export_blocs.pdf",
            #     mime="application/pdf"
            # )

            #Export en PNG
            with st.spinner("Préparation de l'export PNG…"):
                zip_png = export_png_and_zip(code_grid, result_img, alpha_arr, rgb_to_code, block_size)

            st.download_button(
                label="Télécharger le ZIP PNG (image + blocs)",
                data=zip_png,
                file_name="perler_blocs.zip",
                mime="application/zip"
            )

else:
    st.info("Charge une image (JPG ou PNG)")
